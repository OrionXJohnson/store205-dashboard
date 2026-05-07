"""
excel_reader.py

Shared Excel workbook utilities for the Store 205 dashboard project.

Purpose:
- Load workbook files safely.
- Validate required sheets exist.
- Provide reusable workbook helper functions.

Design goals:
- Keep import scripts small and maintainable.
- Centralize workbook validation logic.
- Prevent duplicated Excel loading code.
"""

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIRECTORY = PROJECT_ROOT / "data"


def load_excel_workbook(file_name: str) -> Workbook:
    """
    Load an Excel workbook from the data directory.

    Args:
        file_name:
            Name of the Excel file.

    Returns:
        Loaded openpyxl Workbook object.

    Raises:
        FileNotFoundError:
            If the workbook does not exist.
    """
    workbook_path = DATA_DIRECTORY / file_name

    if not workbook_path.exists():
        raise FileNotFoundError(
            f"Workbook not found: {workbook_path}"
        )

    workbook = load_workbook(
        filename=workbook_path,
        data_only=True
    )

    return workbook


def validate_required_sheets(
    workbook: Workbook,
    required_sheets: list[str]
) -> None:
    """
    Validate that required workbook sheets exist.

    Args:
        workbook:
            Loaded workbook object.

        required_sheets:
            List of required sheet names.

    Raises:
        ValueError:
            If one or more required sheets are missing.
    """
    workbook_sheets = workbook.sheetnames

    missing_sheets = [
        sheet_name
        for sheet_name in required_sheets
        if sheet_name not in workbook_sheets
    ]

    if missing_sheets:
        raise ValueError(
            "Workbook is missing required sheets: "
            f"{missing_sheets}"
        )


def get_sheet_names(workbook: Workbook) -> list[str]:
    """
    Return all workbook sheet names.

    Args:
        workbook:
            Loaded workbook object.

    Returns:
        List of worksheet names.
    """
    return workbook.sheetnames